# -*- coding: utf-8 -*-
"""
Intelligent Enterprise Expense Management & Analytics System
Academic Final Year Project - Enhanced Executive Financial Reporting Suite
"""

import os
import csv
import logging
from datetime import datetime
import matplotlib.pyplot as plt

# External library imports for high-fidelity compiles
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def generate_temp_pdf_chart(expenses, filepath, colors_palette):
    """
    Generates a high-quality category distribution chart and saves it locally
    for embedding inside the ReportLab PDF document story.
    """
    try:
        import collections
        cat_sums = collections.defaultdict(float)
        for exp in expenses:
            cat_sums[exp["category"]] += float(exp["amount"])

        labels = list(cat_sums.keys())
        sizes = list(cat_sums.values())

        if not sizes:
            return False

        fig, ax = plt.subplots(figsize=(6, 3.2), dpi=150)
        fig.patch.set_facecolor("#FFFFFF")
        ax.set_facecolor("#FFFFFF")

        pie_colors = ["#0A84FF", "#30D158", "#FFD60A", "#FF453A", "#8E8E93"]
        
        wedges, texts, autotexts = ax.pie(
            sizes, 
            labels=labels, 
            autopct="%1.1f%%",
            startangle=140,
            colors=pie_colors[:len(sizes)],
            textprops={"color": "#1C1C1E", "fontsize": 8},
            wedgeprops={"edgecolor": "#FFFFFF", "linewidth": 1.5, "antialiased": True}
        )
        for autotext in autotexts:
            autotext.set_color("#FFFFFF")
            autotext.set_weight("bold")

        ax.axis("equal")
        ax.set_title("Outflow Distribution by Channel Category", fontsize=10, fontweight="bold", pad=8)
        fig.tight_layout()

        # Save to temp PNG
        fig.savefig(filepath, facecolor=fig.get_facecolor(), edgecolor='none', bbox_inches='tight')
        plt.close(fig)
        return True
    except Exception as chart_err:
        logging.error("Failed to generate PDF chart: %s", str(chart_err))
        return False


def compile_pdf_report(filepath, expenses, budget_limit, warning_limit, predictor_data, goals_data, currency_symbol="₹"):
    """
    Feature 20: Compiles a professional multi-page ReportLab PDF including:
    - Elegant Cover Page
    - Executive Summary
    - Budget Adherence Review
    - Dynamic Visual Pie Chart
    - Multi-Factor Financial Health Scores
    - Spend Predictions (Monthly, Yearly)
    - Savings Goals Progress Indicators
    - Detailed Transaction Ledger Table
    """
    try:
        # Guarantee parent folder existence
        parent_dir = os.path.dirname(filepath)
        if not os.path.exists(parent_dir):
            os.makedirs(parent_dir)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=letter,
            rightMargin=45, leftMargin=45, topMargin=45, bottomMargin=45
        )
        story = []
        styles = getSampleStyleSheet()

        # Custom typography classes
        title_style = ParagraphStyle(
            "CoverTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=26,
            leading=32,
            textColor=colors.HexColor("#0A84FF"),
            spaceAfter=10
        )
        
        subtitle_style = ParagraphStyle(
            "CoverSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=12,
            textColor=colors.HexColor("#555555"),
            spaceAfter=30
        )

        section_heading = ParagraphStyle(
            "SectionHeading",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#1C1C1E"),
            spaceBefore=15,
            spaceAfter=8,
            keepWithNext=True
        )

        body_style = ParagraphStyle(
            "BodyText",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=13,
            textColor=colors.HexColor("#2C2C2E")
        )

        table_header_style = ParagraphStyle(
            "TableHeader",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=11,
            textColor=colors.white
        )

        # ------------------------------------------------------------------
        # PAGE 1: PROFESSIONAL COVER PAGE
        # ------------------------------------------------------------------
        story.append(Spacer(1, 100))
        story.append(Paragraph("NEURASPEND ENTERPRISE FINANCIAL INTEGRITY REPORT", title_style))
        story.append(Paragraph("Systematic Outflow Audits, Forecast Metrics, and Goals Progress", subtitle_style))
        story.append(Spacer(1, 50))
        
        metadata_data = [
            [Paragraph("<b>Prepared For:</b>", body_style), Paragraph("Academic Viva Board & Capstone Evaluations", body_style)],
            [Paragraph("<b>Prepared By:</b>", body_style), Paragraph("Enterprise Financial Intelligence System", body_style)],
            [Paragraph("<b>Execution Time:</b>", body_style), Paragraph(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), body_style)],
            [Paragraph("<b>Verification Level:</b>", body_style), Paragraph("Administrative Production Grade (Standard PEP-8)", body_style)]
        ]
        meta_table = Table(metadata_data, colWidths=[120, 300])
        meta_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("PADDING", (0, 0), (-1, -1), 6),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5EA")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8F9FA")),
        ]))
        story.append(meta_table)
        story.append(PageBreak())

        # ------------------------------------------------------------------
        # PAGE 2: EXECUTIVE SUMMARY & FINANCIAL VISUALS
        # ------------------------------------------------------------------
        story.append(Paragraph("1. Executive Summary & Forecast Metrics", section_heading))
        story.append(Paragraph("This report presents a thorough offline audit of transactional flows, active budget constraints, and saving schedules compiled on the local database persister.", body_style))
        story.append(Spacer(1, 15))

        # Metrics aggregates
        total_spend = sum(float(exp["amount"]) for exp in expenses)
        remaining = budget_limit - total_spend
        health_score = predictor_data.get("health_score", 100)
        health_rating = predictor_data.get("health_rating", "Excellent")
        next_month_pred = predictor_data.get("predicted_next_month", 0.0)
        yearly_pred = predictor_data.get("predicted_yearly", 0.0)
        trend = predictor_data.get("spending_trend", "Stable")
        highest_growth = predictor_data.get("highest_growth_category", "N/A")

        summary_data = [
            [
                Paragraph("<b>Total Monthly Target:</b>", body_style), Paragraph(f"{currency_symbol}{budget_limit:,.2f}", body_style),
                Paragraph("<b>Forecast Next Month:</b>", body_style), Paragraph(f"{currency_symbol}{next_month_pred:,.2f}", body_style)
            ],
            [
                Paragraph("<b>Cumulative Outflow:</b>", body_style), Paragraph(f"{currency_symbol}{total_spend:,.2f}", body_style),
                Paragraph("<b>Forecast Yearly Spend:</b>", body_style), Paragraph(f"{currency_symbol}{yearly_pred:,.2f}", body_style)
            ],
            [
                Paragraph("<b>Remaining Balance:</b>", body_style), Paragraph(f"{currency_symbol}{remaining:,.2f}", body_style),
                Paragraph("<b>Spending Trend:</b>", body_style), Paragraph(trend, body_style)
            ],
            [
                Paragraph("<b>Financial Health Score:</b>", body_style), Paragraph(f"<b>{health_score}/100</b> ({health_rating})", body_style),
                Paragraph("<b>Escalating Category:</b>", body_style), Paragraph(highest_growth, body_style)
            ]
        ]
        summary_table = Table(summary_data, colWidths=[130, 130, 130, 130])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8F9FA")),
            ("PADDING", (0, 0), (-1, -1), 8),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5EA")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E5EA")),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))

        # Render Pie Chart directly into PDF story
        temp_chart_path = os.path.join(parent_dir, "temp_chart_report.png")
        if generate_temp_pdf_chart(expenses, temp_chart_path, None):
            story.append(Paragraph("Category Distribution Visual Mapping", section_heading))
            story.append(Image(temp_chart_path, width=420, height=224))
            story.append(Spacer(1, 15))

        story.append(PageBreak())

        # ------------------------------------------------------------------
        # PAGE 3: SAVINGS GOALS & DETAILED TRANSACTIONS LEDGER
        # ------------------------------------------------------------------
        story.append(Paragraph("2. Savings Goals Tracking Summary", section_heading))
        if goals_data:
            goals_headers = [Paragraph("<b>Goal Name</b>", table_header_style), Paragraph("<b>Target Target</b>", table_header_style), Paragraph("<b>Amount Saved</b>", table_header_style), Paragraph("<b>Progress Rate</b>", table_header_style)]
            goals_table_data = [goals_headers]
            for g in goals_data:
                saved = float(g["saved"])
                target = float(g["target"])
                pct = (saved / target * 100) if target > 0 else 0.0
                goals_table_data.append([
                    Paragraph(g["title"], body_style),
                    Paragraph(f"{currency_symbol}{target:,.2f}", body_style),
                    Paragraph(f"{currency_symbol}{saved:,.2f}", body_style),
                    Paragraph(f"<b>{pct:.1f}%</b>", body_style)
                ])
            goals_pdf_table = Table(goals_table_data, colWidths=[150, 120, 120, 130])
            goals_pdf_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A84FF")),
                ("PADDING", (0, 0), (-1, -1), 6),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5EA")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E5EA")),
            ]))
            story.append(goals_pdf_table)
        else:
            story.append(Paragraph("No active savings goals committed inside local targets database.", body_style))
        story.append(Spacer(1, 20))

        story.append(Paragraph("3. Detailed Transactions Ledger", section_heading))
        ledger_headers = [
            Paragraph("<b>Date</b>", table_header_style),
            Paragraph("<b>Title</b>", table_header_style),
            Paragraph("<b>Category</b>", table_header_style),
            Paragraph("<b>Payment</b>", table_header_style),
            Paragraph("<b>Amount</b>", table_header_style),
        ]
        
        ledger_data = [ledger_headers]
        # Sort transaction date descending, limited to top 15 for neat page spacing
        sorted_expenses = sorted(expenses, key=lambda x: x["transaction_date"], reverse=True)[:15]
        
        for exp in sorted_expenses:
            ledger_data.append([
                Paragraph(exp["transaction_date"], body_style),
                Paragraph(exp["title"], body_style),
                Paragraph(exp["category"], body_style),
                Paragraph(exp["payment_method"], body_style),
                Paragraph(f"{currency_symbol}{float(exp['amount']):,.2f}", body_style),
            ])

        ledger_table = Table(ledger_data, colWidths=[75, 150, 110, 95, 90])
        ledger_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1C1C1E")),
            ("PADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5EA")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E5EA")),
        ]))
        story.append(ledger_table)

        doc.build(story)
        
        # Clean up temporary chart file safely
        if os.path.exists(temp_chart_path):
            try:
                os.remove(temp_chart_path)
            except Exception:
                pass
        return True
    except Exception as pdf_err:
        logging.error("PDF compiling operation failed: %s", str(pdf_err))
        return False


def compile_monthly_summary_pdf(filepath, expenses, budget_limit, predictor_data, currency_symbol="₹"):
    """
    Feature 19: Generates a short, single-page professional summary PDF including:
    - Title Header
    - Cumulative spends and budget utilization %
    - Top Category and Highest single expense
    - Forecast predictions
    - Financial Health Score
    """
    try:
        parent_dir = os.path.dirname(filepath)
        if not os.path.exists(parent_dir):
            os.makedirs(parent_dir)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=letter,
            rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
        )
        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "SummaryTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=colors.HexColor("#0A84FF"),
            spaceAfter=4
        )
        
        subtitle_style = ParagraphStyle(
            "SummarySubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9.5,
            textColor=colors.HexColor("#555555"),
            spaceAfter=25
        )

        section_heading = ParagraphStyle(
            "SectionHeading",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#1C1C1E"),
            spaceBefore=12,
            spaceAfter=8
        )

        body_style = ParagraphStyle(
            "BodyText",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#2C2C2E")
        )

        story.append(Paragraph("NeuraSpend: Monthly Financial Summary", title_style))
        story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Local Database Audit", subtitle_style))
        story.append(Spacer(1, 10))

        # Core mathematical calculations
        total_spend = sum(float(exp["amount"]) for exp in expenses)
        utilization = (total_spend / budget_limit * 100) if budget_limit > 0 else 0.0
        
        cat_sums = {}
        for exp in expenses:
            cat_sums[exp["category"]] = cat_sums.get(exp["category"], 0.0) + float(exp["amount"])
        top_cat = max(cat_sums, key=cat_sums.get) if cat_sums else "N/A"
        highest_expense = max(float(exp["amount"]) for exp in expenses) if expenses else 0.0

        summary_data = [
            [Paragraph("<b>Total Outflow Spending:</b>", body_style), Paragraph(f"{currency_symbol}{total_spend:,.2f}", body_style)],
            [Paragraph("<b>Budget Utilization Rate:</b>", body_style), Paragraph(f"{utilization:.1f}% of {currency_symbol}{budget_limit:,.2f}", body_style)],
            [Paragraph("<b>Top Expenditure Channel:</b>", body_style), Paragraph(top_cat, body_style)],
            [Paragraph("<b>Peak Single Outflow:</b>", body_style), Paragraph(f"{currency_symbol}{highest_expense:,.2f}", body_style)],
            [Paragraph("<b>Predicted Next Month Outflow:</b>", body_style), Paragraph(f"{currency_symbol}{predictor_data.get('predicted_next_month', 0.0):,.2f}", body_style)],
            [Paragraph("<b>Predicted Yearly Outflow:</b>", body_style), Paragraph(f"{currency_symbol}{predictor_data.get('predicted_yearly', 0.0):,.2f}", body_style)],
            [Paragraph("<b>System Financial Health Score:</b>", body_style), Paragraph(f"<b>{predictor_data.get('health_score', 100)}/100</b> ({predictor_data.get('health_rating', 'Excellent')})", body_style)]
        ]

        summary_table = Table(summary_data, colWidths=[200, 300])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8F9FA")),
            ("PADDING", (0, 0), (-1, -1), 10),
            ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#E5E5EA")),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E5EA")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(summary_table)

        doc.build(story)
        return True
    except Exception as err:
        logging.error("Monthly summary PDF compilation failed: %s", str(err))
        return False


def compile_excel_report(filepath, expenses, budget_limit, goals_data, currency_symbol="₹"):
    """
    Feature 10: Generates a beautiful multi-sheet openpyxl workbook containing:
    - Sheet 1: Expense Records ledger
    - Sheet 2: Budgets Summary (limit, usage, adherence details)
    - Sheet 3: Goals checklists status
    """
    try:
        parent_dir = os.path.dirname(filepath)
        if not os.path.exists(parent_dir):
            os.makedirs(parent_dir)

        workbook = Workbook()

        # Styles definition (Vibrant corporate blue theme)
        header_fill = PatternFill(start_color="0A84FF", end_color="0A84FF", fill_type="solid")
        zebra_fill = PatternFill(start_color="F2F2F7", end_color="F2F2F7", fill_type="solid")
        total_fill = PatternFill(start_color="E5E5EA", end_color="E5E5EA", fill_type="solid")
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=10)
        font_total = Font(name="Calibri", size=11, bold=True)
        
        border_thin = Border(
            left=Side(style="thin", color="D1D1D6"),
            right=Side(style="thin", color="D1D1D6"),
            top=Side(style="thin", color="D1D1D6"),
            bottom=Side(style="thin", color="D1D1D6")
        )
        border_total = Border(
            top=Side(style="thin", color="8E8E93"),
            bottom=Side(style="double", color="1C1C1E")
        )

        # ------------------------------------------------------------------
        # SHEET 1: EXPENSE RECORDS
        # ------------------------------------------------------------------
        sheet1 = workbook.active
        sheet1.title = "Expense Records"

        headers1 = ["ID", "Title", "Description", "Category", "Payment Method", "Amount", "Transaction Date"]
        sheet1.append(headers1)

        # Style sheet 1 headers
        for col_idx in range(1, len(headers1) + 1):
            cell = sheet1.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_thin

        row_num = 2
        for exp in expenses:
            row_data = [
                exp.get("id", ""),
                exp.get("title", ""),
                exp.get("description", ""),
                exp.get("category", ""),
                exp.get("payment_method", ""),
                float(exp.get("amount", 0.0)),
                exp.get("transaction_date", "")
            ]
            sheet1.append(row_data)
            
            fill_style = zebra_fill if row_num % 2 == 0 else None
            for col_idx in range(1, len(row_data) + 1):
                cell = sheet1.cell(row=row_num, column=col_idx)
                cell.font = font_data
                cell.border = border_thin
                if fill_style:
                    cell.fill = fill_style
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx == 6:
                    cell.number_format = f'"{currency_symbol}"#,##0.00'
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx == 7:
                    cell.alignment = Alignment(horizontal="center")
            row_num += 1

        # Sheet 1 Total row
        sheet1.cell(row=row_num, column=5, value="Total Outflow").font = font_total
        sheet1.cell(row=row_num, column=5).alignment = Alignment(horizontal="right")
        sheet1.cell(row=row_num, column=5).fill = total_fill
        
        total_cell = sheet1.cell(row=row_num, column=6)
        total_cell.value = f"=SUM(F2:F{row_num-1})"
        total_cell.font = font_total
        total_cell.number_format = f'"{currency_symbol}"#,##0.00'
        total_cell.alignment = Alignment(horizontal="right")
        total_cell.fill = total_fill
        total_cell.border = border_total

        # Set column width auto-fit for Sheet 1
        for col in sheet1.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet1.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # ------------------------------------------------------------------
        # SHEET 2: BUDGETS SUMMARY
        # ------------------------------------------------------------------
        sheet2 = workbook.create_sheet(title="Budgets Summary")
        headers2 = ["Parameter", "Target Value"]
        sheet2.append(headers2)

        for col_idx in range(1, len(headers2) + 1):
            cell = sheet2.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_thin

        total_spend = sum(float(exp["amount"]) for exp in expenses)
        remaining = budget_limit - total_spend
        usage_pct = (total_spend / budget_limit) if budget_limit > 0 else 0.0

        budget_summary_data = [
            ["Monthly Budget Target Limit", budget_limit],
            ["Cumulative Spending Volume", total_spend],
            ["Remaining Balance Available", remaining],
            ["Budget Utilization rate", usage_pct]
        ]

        row_idx = 2
        for key, val in budget_summary_data:
            sheet2.cell(row=row_idx, column=1, value=key).font = font_data
            sheet2.cell(row=row_idx, column=1).border = border_thin
            
            val_cell = sheet2.cell(row=row_idx, column=2, value=val)
            val_cell.font = font_data
            val_cell.border = border_thin
            if row_idx == 5:
                val_cell.number_format = "0.0%"
                val_cell.alignment = Alignment(horizontal="right")
            else:
                val_cell.number_format = f'"{currency_symbol}"#,##0.00'
                val_cell.alignment = Alignment(horizontal="right")
            row_idx += 1

        sheet2.column_dimensions["A"].width = 30
        sheet2.column_dimensions["B"].width = 18

        # ------------------------------------------------------------------
        # SHEET 3: GOALS CHECKLISTS
        # ------------------------------------------------------------------
        sheet3 = workbook.create_sheet(title="Goals Reports")
        headers3 = ["Goal Name", "Target Amount", "Saved Amount", "Completion Rate"]
        sheet3.append(headers3)

        for col_idx in range(1, len(headers3) + 1):
            cell = sheet3.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_thin

        row_idx = 2
        for g in goals_data:
            saved = float(g["saved"])
            target = float(g["target"])
            rate = (saved / target) if target > 0 else 0.0

            sheet3.cell(row=row_idx, column=1, value=g["title"]).font = font_data
            sheet3.cell(row=row_idx, column=1).border = border_thin

            t_cell = sheet3.cell(row=row_idx, column=2, value=target)
            t_cell.font = font_data
            t_cell.number_format = f'"{currency_symbol}"#,##0.00'
            t_cell.alignment = Alignment(horizontal="right")
            t_cell.border = border_thin

            s_cell = sheet3.cell(row=row_idx, column=3, value=saved)
            s_cell.font = font_data
            s_cell.number_format = f'"{currency_symbol}"#,##0.00'
            s_cell.alignment = Alignment(horizontal="right")
            s_cell.border = border_thin

            r_cell = sheet3.cell(row=row_idx, column=4, value=rate)
            r_cell.font = font_data
            r_cell.number_format = "0.0%"
            r_cell.alignment = Alignment(horizontal="right")
            r_cell.border = border_thin

            row_idx += 1

        sheet3.column_dimensions["A"].width = 25
        sheet3.column_dimensions["B"].width = 18
        sheet3.column_dimensions["C"].width = 18
        sheet3.column_dimensions["D"].width = 18

        workbook.save(filepath)
        return True
    except Exception as excel_err:
        logging.error("Excel compile failed: %s", str(excel_err))
        return False
